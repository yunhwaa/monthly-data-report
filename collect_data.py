import os
import smtplib
import pandas as pd
import yfinance as yf
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from PublicDataReader import Kosis

# ── 설정 ──────────────────────────────────────────────
KOSIS_KEY   = os.environ["KOSIS_API_KEY"]
GMAIL_USER  = os.environ["GMAIL_USER"]
GMAIL_PASS  = os.environ["GMAIL_APP_PASSWORD"]
TO_EMAIL    = "kyhh9124@naver.com"

NOW         = datetime.now()
YEAR        = NOW.strftime("%Y")
PREV_MONTH  = (NOW.replace(day=1) - pd.DateOffset(months=1))
START_PRD   = PREV_MONTH.strftime("%Y%m")
END_PRD     = START_PRD
REPORT_NAME = f"monthly_report_{NOW.strftime('%Y%m')}.xlsx"
# ──────────────────────────────────────────────────────


def collect_kosis() -> pd.DataFrame:
    try:
        api = Kosis(KOSIS_KEY)
        df = api.get_data(
            "통계자료",
            orgId      = "134",
            tblId      = "DT_134001_001",
            itmId      = "ALL",
            objL1      = "ALL",
            prdSe      = "M",
            startPrdDe = START_PRD,
            endPrdDe   = END_PRD,
        )
        if df is None or df.empty:
            print(f"⚠️  KOSIS 데이터 없음 (기간: {START_PRD}~{END_PRD}) → 빈 시트로 대체")
            return pd.DataFrame({"안내": [f"{START_PRD} 기간 데이터가 없습니다."]})
        print(f"✅ KOSIS 데이터 {len(df)}행 수집 완료")
        return df
    except Exception as e:
        print(f"⚠️  KOSIS 수집 오류: {e} → 빈 시트로 대체")
        return pd.DataFrame({"오류": [str(e)]})


def collect_yfinance() -> tuple[pd.DataFrame, pd.DataFrame]:
    start = f"{YEAR}-01-01"
    end   = NOW.strftime("%Y-%m-%d")

    dow    = yf.download("^DJI",  start=start, end=end, interval="1mo", progress=False)
    nasdaq = yf.download("^IXIC", start=start, end=end, interval="1mo", progress=False)

    # yfinance 버전에 따라 컬럼 구조가 다를 수 있어서 안전하게 처리
    if isinstance(dow.columns, pd.MultiIndex):
        dj_close = dow["Close"].iloc[:, 0]
        nq_close = nasdaq["Close"].iloc[:, 0]
    else:
        dj_close = dow["Close"]
        nq_close = nasdaq["Close"]

    dj_df = dj_close.reset_index().rename(columns={dj_close.name: "Dow Jones Close", "index": "Date"})
    nq_df = nq_close.reset_index().rename(columns={nq_close.name: "NASDAQ Close", "index": "Date"})

    # Date 컬럼 포맷 정리
    for df in [dj_df, nq_df]:
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"]).dt.strftime("%Y-%m")

    print(f"✅ Yahoo Finance 데이터 수집 완료 (DJI: {len(dj_df)}행, NASDAQ: {len(nq_df)}행)")
    return dj_df, nq_df


def style_sheet(writer: pd.ExcelWriter, sheet_name: str):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = writer.sheets[sheet_name]
    header_fill = PatternFill("solid", start_color="1F4E79")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def save_excel(kosis_df: pd.DataFrame, dj_df: pd.DataFrame, nq_df: pd.DataFrame) -> str:
    path = f"/tmp/{REPORT_NAME}"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        kosis_df.to_excel(writer, sheet_name="KOSIS_미분양현황", index=False)
        dj_df.to_excel(writer,    sheet_name="DowJones",         index=False)
        nq_df.to_excel(writer,    sheet_name="NASDAQ",           index=False)
        style_sheet(writer, "KOSIS_미분양현황")
        style_sheet(writer, "DowJones")
        style_sheet(writer, "NASDAQ")
    print(f"✅ 엑셀 저장 완료: {REPORT_NAME}")
    return path


def send_email(file_path: str):
    subject = f"[월간 자동 리포트] {NOW.strftime('%Y년 %m월')} 데이터"
    body    = (
        f"{NOW.strftime('%Y년 %m월')} 월간 데이터 리포트입니다.\n\n"
        "포함 시트:\n"
        "  1. KOSIS_미분양현황 – KOSIS 통계자료 (134 / DT_134001_001)\n"
        "  2. DowJones         – 다우존스 월봉 종가\n"
        "  3. NASDAQ           – 나스닥 월봉 종가\n\n"
        "자동 발송 메일입니다."
    )

    msg = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = TO_EMAIL
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(file_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{REPORT_NAME}"')
        msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_PASS)
        server.sendmail(GMAIL_USER, TO_EMAIL, msg.as_string())

    print(f"✅ 이메일 발송 완료 → {TO_EMAIL}")


def main():
    print(f"🗓  수집 기간: {START_PRD}")

    print("📊 KOSIS 데이터 수집 중...")
    kosis_df = collect_kosis()

    print("📈 Yahoo Finance 데이터 수집 중...")
    dj_df, nq_df = collect_yfinance()

    print("💾 엑셀 파일 저장 중...")
    file_path = save_excel(kosis_df, dj_df, nq_df)

    print("📧 이메일 발송 중...")
    send_email(file_path)


if __name__ == "__main__":
    main()
